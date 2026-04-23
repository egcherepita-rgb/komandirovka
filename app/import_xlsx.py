from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from .database import Base, SessionLocal, engine
from .main import make_slot_code
from .models import Slot


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def run_import(xlsx_path: str) -> None:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f'Файл не найден: {path}')

    Base.metadata.create_all(bind=engine)
    wb = load_workbook(path, data_only=True)
    ws = wb['Бронь_слоты']

    db = SessionLocal()
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            week, period, direction, status, branch, contact, goal, priority, comment, slot_code = row[:10]
            if week is None or direction is None:
                continue
            week = int(week)
            direction = str(direction).strip()
            existing = db.execute(
                select(Slot).where(Slot.week == week, Slot.product_direction == direction)
            ).scalar_one_or_none()
            if existing:
                continue

            db.add(
                Slot(
                    week=week,
                    period=str(period or ''),
                    product_direction=direction,
                    status=str(status or 'Свободно'),
                    branch=clean(branch),
                    contact=clean(contact),
                    visit_goal=clean(goal),
                    priority=clean(priority),
                    comment=clean(comment),
                    slot_code=str(slot_code or make_slot_code(week, direction)),
                )
            )
        db.commit()
    finally:
        db.close()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Использование: python -m app.import_xlsx "путь_к_файлу.xlsx"')
        raise SystemExit(1)
    run_import(sys.argv[1])
    print('Импорт завершён')
