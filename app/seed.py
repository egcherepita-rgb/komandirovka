from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import Slot


DEFAULT_STATUSES = ['Свободно', 'Предзапрос', 'Забронировано', 'Отменено', 'Отпуск']
DEFAULT_PRIORITIES = ['Высокий', 'Средний', 'Низкий']


def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def get_seed_path() -> Path:
    raw = os.getenv('SEED_XLSX', './data/Календарь_командировок_2026 (1).xlsx')
    return Path(raw)


def get_reference_lists(seed_path: Path) -> dict:
    result = {
        'statuses': DEFAULT_STATUSES,
        'priorities': DEFAULT_PRIORITIES,
        'branches': [],
        'directions': [],
    }
    if not seed_path.exists():
        return result

    wb = load_workbook(seed_path, data_only=True)
    if 'Справочники' not in wb.sheetnames:
        return result

    ws = wb['Справочники']
    branches, directions, statuses, priorities = [], [], [], []

    row = 3
    while row <= ws.max_row:
        branch = _clean(ws[f'A{row}'].value)
        direction = _clean(ws[f'B{row}'].value)
        status = _clean(ws[f'J{row}'].value)
        priority = _clean(ws[f'L{row}'].value)
        if branch:
            branches.append(branch)
        if direction:
            directions.append(direction)
        if status:
            statuses.append(status)
        if priority:
            priorities.append(priority)
        row += 1

    if branches:
        result['branches'] = list(dict.fromkeys(branches))
    if directions:
        result['directions'] = list(dict.fromkeys(directions))
    if statuses:
        result['statuses'] = list(dict.fromkeys(statuses))
    if priorities:
        result['priorities'] = list(dict.fromkeys(priorities))
    return result


def seed_slots_if_empty(db: Session, seed_path: Path) -> None:
    existing = db.scalar(select(Slot.id).limit(1))
    if existing:
        return
    if not seed_path.exists():
        return

    wb = load_workbook(seed_path, data_only=True)
    ws = wb['Бронь_слоты']

    for row in ws.iter_rows(min_row=2, values_only=True):
        week, period, direction, status, branch, contact, goal, priority, comment, slot_code = row[:10]
        if week is None or direction is None or slot_code is None:
            continue
        db.add(
            Slot(
                week=int(week),
                period=str(period),
                product_direction=str(direction),
                status=str(status or 'Свободно'),
                branch=_clean(branch),
                contact=_clean(contact),
                visit_goal=_clean(goal),
                priority=_clean(priority) or 'Средний',
                comment=_clean(comment),
                slot_code=str(slot_code),
            )
        )
    db.commit()
